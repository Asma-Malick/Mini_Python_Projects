import openpyxl as xl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import sys

def process_workbook(filename, chart_type='BarChart'):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    categories = Reference(sheet,
                           min_row=2,
                           max_row=sheet.max_row,
                           min_col=2,
                           max_col=2)
    
    if chart_type == 'BarChart':
        chart = BarChart()
    elif chart_type == 'LineChart':
        chart = LineChart()
    elif chart_type == 'PieChart':
        chart = PieChart()
    else:
        print("Invalid chart type. Choose from 'BarChart', 'LineChart', 'PieChart'.")
        sys.exit(1)

    chart.add_data(values, titles_from_data=True)
    chart.set_categories(categories)
    chart.title = "Corrected Prices"
    
    if hasattr(chart, 'y_axis'):
        chart.y_axis.title = 'Price'
    if hasattr(chart, 'x_axis'):
        chart.x_axis.title = 'Product ID'
    
    sheet.add_chart(chart, 'E2')

    wb.save('transaction3.xlsx')

if __name__ == "__main__":
    filename = input("Enter the filename: ")
    chart_type = input("Enter chart type (BarChart, LineChart, PieChart): ")
    process_workbook(filename, chart_type)
